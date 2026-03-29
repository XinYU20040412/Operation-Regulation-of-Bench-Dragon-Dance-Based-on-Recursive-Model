# -*- coding: utf-8 -*-
"""
Created on Sat Sep  7 20:05:49 2024

@author: ASUS
"""

from math import pi,sqrt,cos,log,sin,atan,tan
import numpy as np
from scipy.optimize import fsolve

def head_position_per_second():
    #求解每一时刻龙头位置
    k=1.7/(2*pi)
    C=0.5*k**2*log(k)
    def theta(r):#定义两个转换工具
        return r/k
    def r_(theta_):
        return k*theta_
    rB = 4.5
    tB = (0.5 * (rB * sqrt(rB**2 + k**2) + k**2 * log(rB + sqrt(rB**2 + k**2))) - C) /k 
    rA = r_(32 * pi)
    tA = (0.5 * (rA * sqrt(rA**2 + k**2) + k**2 * log(rA + sqrt(rA**2 + k**2))) - C) / k
    t0=tB+100
    tf=tB


    t = [i for i in np.arange(t0,tf-1,-1 )]
    rr = []
    x = []
    y = []
    theta_ = []

    for i in t:
        def equation(r, t=i):
            return 0.5 * (r * sqrt(r**2 + k**2) + k**2 * log(r + sqrt(r**2 + k**2))) - k * t - C

        initial_guess = 1
        solution = fsolve(equation, initial_guess)
        r = solution[0]
        rr.append(r)
        a = theta(r)
        theta_.append(a)
        x0 = r * cos(a)
        y0 = r * sin(a)
        x.append(x0)
        y.append(y0)
    return x,y

class constant(object):
    def __init__(self):
        self.num_chair=223
        self.d_head=0.01*(341-2*27.5)
        self.d_gen=0.01*(220-2*27.5)

constant=constant()

def xy_to_polar(x,y):
    r0=1.7
    r=np.sqrt(x**2+y**2)
    theta=2*pi*(r/r0)
    return r,theta

#已知每个时刻的龙头位置x,y(x_head,y_head)
#迭代二分法找对应点
def point_equation(delta_theta,x1,y1,r0,d):
    #选取点与上一位置点的距离差
    _,theta1=xy_to_polar(x1,y1)
    theta2=theta1+delta_theta
    x2=r0*theta2*cos(theta2)/(2*pi)
    y2=r0*theta2*sin(theta2)/(2*pi)
    distance=(x1-x2)**2+(y1-y2)**2
    return distance-d**2

def count_position(x_head,y_head,constant):
    #得出每一点的位置(x,y,r)
    xs,ys,rs,r0=[x_head],[y_head],[sqrt(x_head**2+y_head**2)],1.7
    positions_moment=[x_head,y_head]
    for i in range(constant.num_chair):
        x,y=xs[i],ys[i]
        if i==0:
            delta_theta=fsolve(point_equation, pi/4,args=(x,y,r0,constant.d_head))
        else:
            delta_theta=fsolve(point_equation, pi/4,args=(x,y,r0,constant.d_gen))
        assert delta_theta>0, "Value must be positive"
        _,theta1=xy_to_polar(x,y)
        theta2=theta1+delta_theta
        r_new=theta2*r0/(2*pi)
        if r_new>16*1.7:
            break
        x_new=r0*theta2*cos(theta2)/(2*pi)
        y_new=r0*theta2*sin(theta2)/(2*pi)
        xs.append(x_new)
        ys.append(y_new)
        positions_moment.append(x_new)
        positions_moment.append(y_new)
        rs.append(r_new)
    while len(ys)<constant.num_chair:
        i=1
        if i==1:
            if len(ys)==1:
                y_new=sqrt(constant.d_head**2-(16*1.7-xs[-1])**2)+ys[-1]
            else:
                y_new=sqrt(constant.d_gen**2-(16*1.7-xs[-1])**2)+ys[-1]
        else:
            y_new=ys[-1]+constant.d_gen
        xs.append(16*1.7)
        ys.append(y_new)
        positions_moment.append(16*1.7)
        positions_moment.append(y_new)
        rs.append(sqrt((16*1.7)**2+y_new**2))
        i+=1
    return xs,ys,positions_moment,rs

def count_angle(xs,ys,rs,constant):
    #计算每条凳子前端与后端与相应切线或者竖直直线的夹角
    def linear_para(x1,y1,x2,y2):#得出一般直线方程系数
        return y1-y2,x2-x1
    angles_front,angles_behind,r0=[],[],1.7
    for i in range(constant.num_chair):
        if xs[i]==16*1.7:
            angles_front.append(0)
            angles_behind.append(0)
        elif xs[i]!=16*1.7 and xs[i+1]==16*1.7:
            alpha1=atan(r0/(2*pi*rs[i]))
            A,B=linear_para(xs[i],ys[i],xs[i+1],ys[i+1])
            beta=atan((A*tan(2*pi*rs[i]/r0)-B)/(A+B*tan(2*pi*rs[i]/r0)))
            angles_front.append(alpha1-beta)
            angles_behind.append(atan(B/A))
        else:
            alpha1=atan(r0/(2*pi*rs[i]))
            alpha2=atan(r0/(2*pi*rs[i+1]))
            A,B=linear_para(xs[i],ys[i],xs[i+1],ys[i+1])
            beta=atan((A*tan(2*pi*rs[i]/r0)-B)/(A+B*tan(2*pi*rs[i]/r0)))
            angles_front.append(abs(alpha1-beta))
            angles_behind.append(abs(alpha2-beta))
    return angles_front,angles_behind

def count_velocity(angles_front,angles_behind):
    velocitys=[1]
    for i in range(constant.num_chair):
        velocity_new=velocitys[i]*cos(angles_front[i])/cos(angles_behind[i])
        velocitys.append(velocity_new)
    return velocitys
        
time=100
xs_head,ys_head=head_position_per_second()
positions,velocitys=[],[]
for i in range(time+1):
    x_head,y_head=xs_head[i],ys_head[i]
    xs,ys,positions_moment,rs=count_position(x_head, y_head,constant)
    positions.append(positions_moment)
    angles_front,angles_behind=count_angle(xs, ys, rs, constant)
    velocitys_moment=count_velocity(angles_front,angles_behind)
    velocitys.append(velocitys_moment)


def ensure_numeric(lst):
    return [[float(x) if not np.isnan(x) else np.nan for x in sublist] for sublist in lst]
positions = ensure_numeric(positions)
velocitys = ensure_numeric(velocitys)



from openpyxl import load_workbook
positions = np.array(positions)
velocitys = np.array(velocitys)
positions = positions.T
velocitys = velocitys.T
file_path = r"C:\Users\ASUS\Desktop\附件\result4.xlsx"
wb = load_workbook(file_path)
start_row = 2
start_col = 2
ws = wb.active
ws1 = wb['位置']
ws2 = wb['速度']
for i, row in enumerate(positions):
    for j, value in enumerate(row):
        ws1.cell(row=start_row + i, column=start_col + j, value=value)
for i, row in enumerate(velocitys):
    for j, value in enumerate(row):
        ws2.cell(row=start_row + i, column=start_col + j, value=value)
wb.save(file_path)
